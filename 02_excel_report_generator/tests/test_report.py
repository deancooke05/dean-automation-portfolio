import tempfile, unittest
from pathlib import Path
from openpyxl import load_workbook
from src.analytics import analyse, load_sales
from src.report import generate_report

ROOT=Path(__file__).resolve().parents[1]
DATA=ROOT/"sample_data/sample_sales.csv"

class Product002Tests(unittest.TestCase):
    def test_loads_sample(self): self.assertEqual(len(load_sales(DATA)),30)
    def test_metrics(self):
        a=analyse(load_sales(DATA)); self.assertEqual(a.orders,30); self.assertGreater(a.revenue,0); self.assertGreaterEqual(a.highest_sale,a.lowest_sale)
    def test_groups_reconcile(self):
        a=analyse(load_sales(DATA)); self.assertAlmostEqual(a.revenue,float(a.by_product.sum()),2); self.assertAlmostEqual(a.revenue,float(a.by_region.sum()),2)
    def test_generates_workbook(self):
        with tempfile.TemporaryDirectory() as d:
            p=generate_report(DATA,Path(d)/"report.xlsx"); self.assertTrue(p.exists())
            wb=load_workbook(p,data_only=False); self.assertEqual(wb.sheetnames,["Executive Summary","Analysis","Raw Data","Report Information"])
            self.assertEqual(wb["Executive Summary"]["A1"].value,"COOKE AUTOMATION SYSTEMS")
            self.assertEqual(wb["Raw Data"].max_row,31)

if __name__=="__main__": unittest.main()
